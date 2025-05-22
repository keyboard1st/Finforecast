import logging
import os
import time

import json
from datetime import datetime
from dataclasses import dataclass, asdict
import pandas as pd

from openpyxl import load_workbook

def create_logger(logger_file_path):

    if not os.path.exists(logger_file_path):
        os.makedirs(logger_file_path)
    log_name = '{}.log'.format(time.strftime('%Y-%m-%d'))
    final_log_file = os.path.join(logger_file_path, log_name)

    logger = logging.getLogger()  # 设定日志对象
    logger.setLevel(logging.INFO)  # 设定日志等级

    file_handler = logging.FileHandler(final_log_file)  # 文件输出
    console_handler = logging.StreamHandler()  # 控制台输出

    # 输出格式
    formatter = logging.Formatter(
        "%(asctime)s - %(message)s "
    )

    file_handler.setFormatter(formatter)  # 设置文件输出格式
    console_handler.setFormatter(formatter)  # 设施控制台输出格式
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger

@dataclass
class Config:
    """实验配置类（带类型提示和默认值）"""
    config_type: str = 'GRU双层实验'
    model_type: str = 'GRU'
    input_dim: int = 68
    hidden_dim: int = 64
    num_layers: int = 1
    output_dim: int = 1
    dropout: float = 0.0
    learning_rate: float = 1e-4
    shuffle_time: bool = False
    early_stop_patience: int = 20
    train_epochs: int = 200
    window_size: int = 30
    num_val_windows: int = 200
    val_sample_mode : str = "random"
    loss: str = 'MSE'
    pct_start: float = 0.2
    lradj: str = 'TST'
    exp_path: str = '/home/hongkou/chenx/exp/'
    cross_train: bool = False


    @classmethod
    def from_json(cls, file_path: str):
        """从JSON文件创建配置实例"""
        with open(file_path, 'r') as f:
            config_dict = json.load(f)
        return cls(**config_dict)

    def to_json(self, save_dir: str = None):
        """将配置保存为JSON文件"""
        os.makedirs(save_dir or self.model_path, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_path = os.path.join(save_dir or self.model_path, f"config_{timestamp}.json")

        with open(save_path, 'w') as f:
            json.dump(asdict(self), f, indent=4)
        return save_path

    def validate(self):
        """参数合法性验证"""
        assert self.learning_rate > 0, "学习率必须为正数"
        assert self.num_layers >= 1, "网络层数至少为1"
        assert self.window_size > 0, "时间窗口大小必须为正整数"


def record_to_excel(df_config, df_metrics, df_result, save_path, append=False):
    """
    :param append: True=追加模式，False=覆盖模式
    """
    file_path = os.path.join(save_path, 'exp.xlsx')

    if append and os.path.exists(file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # 获取workbook对象
            workbook = writer.book

            # 添加分隔行
            for sheet_name in ['实验配置', '实验过程', '实验结果']:
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    max_row = ws.max_row
                    # 插入分隔行（跳过空表）
                    if max_row > 0:
                        ws.append(['===== 新一轮实验数据 ====='])
                        ws.append([])

            # 写入新数据（跳过标题行）
            df_config.to_excel(writer, sheet_name='实验配置', index=False, header=False,
                               startrow=writer.sheets['实验配置'].max_row)
            df_metrics.to_excel(writer, sheet_name='实验过程', index=False, header=False,
                                startrow=writer.sheets['实验过程'].max_row)
            df_result.to_excel(writer, sheet_name='实验结果', index=False, header=False,
                               startrow=writer.sheets['实验结果'].max_row)
    else:
        # 覆盖模式（首次写入）
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_config.to_excel(writer, sheet_name='实验配置', index=False)
            df_metrics.to_excel(writer, sheet_name='实验过程', index=False)
            df_result.to_excel(writer, sheet_name='实验结果', index=False)

# 使用示例
if __name__ == "__main__":
    exp_path = r'C:\Users\chen.xing\key\GRU\Experiment\exp_001'
    config = Config.from_json(os.path.join(exp_path, 'config.json'))

    # 修改部分参数
    config.train_epochs = 100
    config.hidden_dim = 128

    # 验证参数
    config.validate()
    logger = create_logger(os.path.join(exp_path))

    logger.info(f"Config loaded: {config.__dict__}")